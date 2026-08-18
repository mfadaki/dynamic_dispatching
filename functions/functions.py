import os
import json
import shelve
import logging
from time import time
from datetime import datetime

import numpy as np
from openpyxl import Workbook


class NumpyArrayEncoder(json.JSONEncoder):
    """json.JSONEncoder subclass that knows how to serialize numpy types.

    saveResultsFn's format_json branch calls json.dump(..., cls=NumpyArrayEncoder);
    without this, dumping a numpy array, numpy scalar (e.g. np.float64), or a
    dict/list containing them raises TypeError: Object of type ndarray/float64
    is not JSON serializable.
    """
    def default(self, obj):
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, (np.bool_,)):
            return bool(obj)
        return json.JSONEncoder.default(self, obj)


def dim(vec):
    return np.array(vec).shape


def RunTime(start_time):
    end_time = time()
    run_time_seconds = end_time - start_time  # time_taken is in seconds
    hours, rest = divmod(run_time_seconds, 3600)
    minutes, seconds = divmod(rest, 60)
    hours_str = str(round(hours)) if hours > 9 else "0" + str(round(hours))
    minutes_str = str(round(minutes)) if minutes > 9 else "0" + str(round(minutes))
    seconds_str = str(round(seconds)) if seconds > 9 else "0" + str(round(seconds))
    run_time = hours_str + ":" + minutes_str + ":" + seconds_str
    print("*" * 50)
    print("Run Time: ", run_time)
    print("*" * 50)
    return run_time

def get_object_type(obj):
    if isinstance(obj, type):
        return "class"
    elif not isinstance(obj, (int, float, str, list, dict, set, tuple)) and hasattr(obj, "__class__"):
        return "instance"
    elif isinstance(obj, int):
        return "int"
    elif isinstance(obj, float):
        return "float"
    elif isinstance(obj, str):
        return "str"
    elif isinstance(obj, list):
        return "list"
    elif isinstance(obj, dict):
        return "dict"
    elif isinstance(obj, set):
        return "set"
    elif isinstance(obj, tuple):
        return "tuple"
    else:
        return "unknown"


def json_to_excel(results_json):
    wb = Workbook()
    wb.remove(wb["Sheet"])
    wb.create_sheet("results")
    ws = wb["results"]

    col_headers = results_json[str(0)].keys()

    col_no = 1
    for cl in col_headers:
        ws.cell(row=1, column=col_no).value = cl
        col_no = col_no + 1

    for rw in range(len(results_json.keys())):
        cn = 0
        for cl in col_headers:
            ws.cell(row=rw + 2, column=cn + 1).value = results_json[str(rw)][cl]
            cn = cn + 1

    wb.save("./results/" + nameof(results_json) + ".xlsx")
    return None

#If the keys of a dict in any level (eg. 1,2,3...) would be tuple, we can't save it as json as the keys should be string. This function recursively converts all keys from tuple to string.
def convert_keys_to_strings(d):
    """Recursively converts all tuple keys in a dictionary to strings."""
    if isinstance(d, dict):
        return {str(k): convert_keys_to_strings(v) for k, v in d.items()}
    elif isinstance(d, list):
        return [convert_keys_to_strings(v) for v in d]
    else:
        return d

def saveResultsFn(save_dict, settings_markdown_string, filename):
    date_time_stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    os.makedirs("./results", exist_ok=True)   # os.mkdir below fails if this parent is missing
    os.mkdir("./results/" + date_time_stamp + "_{" + filename + "}/")

    format_shelve = save_dict['format_shelve']
    format_json = save_dict['format_json']
    format_cplex_model = save_dict['format_cplex_model']

    for key, value in format_shelve.items():
        with shelve.open("./results/" + date_time_stamp + "_{" + filename + "}/" + key) as db:
            db[key] = value
        logging.debug( f'saved: {key}')

    for key, value in format_json.items():
        with open("./results/" + date_time_stamp + "_{" + filename + "}/" + key + ".json", "w") as write_file:
            json.dump(value, write_file, cls=NumpyArrayEncoder)
        logging.debug( f'saved: {key}')

    for key, value in format_cplex_model.items():
        value.export_as_sav("./results/" + date_time_stamp + "_{" + filename + "}/" + key + ".sav")
        logging.debug( f'saved: {key}')

    # Saving the Markdown file which includes the Settings of the Experiment
    with open("./results/" + date_time_stamp + "_{" + filename + "}/" + "settings.md", "w") as f:
        f.write(settings_markdown_string)
        logging.debug( f'saved: settings.md')

    return "./results/" + date_time_stamp + "_{" + filename + "}/"

def lastFolder_fn():
    results_folder = "./results/"

    # List all subdirectories with their modification times
    subdirs = [
        (d, os.path.getmtime(os.path.join(results_folder, d)))
        for d in os.listdir(results_folder)
        if os.path.isdir(os.path.join(results_folder, d))
    ]

    # Sort subdirectories by modification time (latest first)
    sorted_subdirs = sorted(subdirs, key=lambda x: x[1], reverse=True)

    # Get the latest folder
    latest_folder = sorted_subdirs[0][0] if sorted_subdirs else None

    if latest_folder:
        print("The latest folder is:", latest_folder)
    else:
        print("No folders found in the directory.")
    return latest_folder