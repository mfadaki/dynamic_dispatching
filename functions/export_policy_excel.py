"""
export_policy_excel.py
======================
Call this from main.py after run_psmd() finishes:

    from export_policy_excel import export_policy_excel
    export_policy_excel(mdp, alp, thetabar, n_days=2, seed=7)

Output: dispatching_policy_report.xlsx  (written to current directory)
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── helpers ──────────────────────────────────────────────────────────────────

def _solid(hex_):
    return PatternFill('solid', fgColor=hex_)

def _border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _hfont(size=9, bold=True, color='FFFFFF'):
    return Font(name='Arial', size=size, bold=bold, color=color)

def _dfont(size=9, bold=False):
    return Font(name='Arial', size=size, bold=bold, color='000000')

def _center():
    return Alignment(horizontal='center', vertical='center')

def _right():
    return Alignment(horizontal='right', vertical='center')

def _left():
    return Alignment(horizontal='left', vertical='center')


# ── simulation ────────────────────────────────────────────────────────────────

def _run_simulation(mdp, alp, thetabar, n_days, seed):
    """Run greedy trajectory; return (rows, end_of_day_exps).

    The MDP state no longer carries an expiry counter or a day index
    (Section 3.3 — infinite-horizon, stationary model): both are derived
    here purely for the human-readable report. Day is tracked as external
    simulation bookkeeping, incremented whenever an end-of-day transition is
    detected. Realised (not expected) expiry is computed on the same
    post-dispatch, post-event inventory the transition function itself uses
    (mdp.apply_action_and_event), with the actual drawn event g — this is
    the one-trajectory realisation of the same quantity mdp.cost() averages
    over all possible events in closed form for the optimisation itself.
    """
    from classes.bounds import UpperBound
    from inputs.inputs import (N_LABS, L_AGE, LAMBDA_TOTAL, C_DISPATCH,
                                H_HOLD, C_EXP_DEPOT, C_EXP_LAB,
                                TAU_MAX, DELTA_T)

    np.random.seed(seed)
    ub = UpperBound(mdp, alp)

    s = mdp.sample_initial_state()
    s[mdp.TAU_IDX] = TAU_MAX

    rows = []
    eod_exps = {}   # day -> {eps_depot, eps_labs}   (report-only bookkeeping)
    day = 0

    for _ in range(int(TAU_MAX / DELTA_T) * n_days + 10):
        if day >= n_days:
            break
        n, tau = mdp.parse_state(s)

        a = ub.greedy_action(s, thetabar, mdp.sample_exog(200))
        g = mdp.sample_single_exog()

        arrive_age  = int(g)         if g < L_AGE                 else None
        process_lab = int(g - L_AGE) if L_AGE <= g < L_AGE+N_LABS else None

        depot_snap = n[0].copy().astype(int)
        labs_snap  = n[1:].copy().astype(int)       # (N_LABS, L_AGE)
        processed  = np.zeros((N_LABS, L_AGE), dtype=int)
        if process_lab is not None:
            for ai in range(L_AGE - 1, -1, -1):
                if labs_snap[process_lab, ai] > 0:
                    processed[process_lab, ai] = 1
                    break

        c_disp = float(C_DISPATCH[a - 1]) if a > 0 else 0.0
        c_hold = float(
            (H_HOLD[0] * n[0]).sum() / LAMBDA_TOTAL
            + sum((H_HOLD[p+1] * n[p+1]).sum() / LAMBDA_TOTAL
                  for p in range(N_LABS))
        )

        # Realised expiry (report-only): same post-dispatch, post-event
        # inventory the transition uses internally, with the actual drawn
        # event g — nonzero only on the epoch that triggers end-of-day.
        is_eod = (tau - DELTA_T <= 0)
        eps_depot_disp, eps_labs_disp, c_exp = 0, [0] * N_LABS, 0.0
        if is_eod:
            n_after = mdp.apply_action_and_event(n.astype(float), a, g)
            eps_depot_disp = int(n_after[0, 0])
            eps_labs_disp  = [int(n_after[p + 1, 0]) for p in range(N_LABS)]
            c_exp = float(C_EXP_DEPOT * eps_depot_disp
                          + C_EXP_LAB * sum(eps_labs_disp))
            eod_exps[day] = dict(eps_depot=eps_depot_disp, eps_labs=eps_labs_disp)

        rows.append(dict(
            step       = len(rows),
            day        = day,
            tau        = round(float(tau), 4),
            event      = ('arrive a' + str(arrive_age  + 1) if arrive_age  is not None else
                          'proc L'   + str(process_lab + 1) if process_lab is not None else
                          'dummy'),
            arrive     = [1 if arrive_age == k else 0 for k in range(L_AGE)],
            depot      = depot_snap.tolist(),
            labs       = labs_snap.tolist(),
            processed  = processed.tolist(),
            action     = int(a),
            eps_depot  = eps_depot_disp,
            eps_labs   = eps_labs_disp,
            c_disp     = c_disp,
            c_hold     = round(c_hold, 6),
            c_exp      = c_exp,
        ))

        s = np.asarray(mdp.transition(s, a, g), dtype=float).ravel()
        if is_eod:
            day += 1

    return rows, eod_exps


# ── sheet builders ────────────────────────────────────────────────────────────

def _build_simulation_sheet(wb, rows, eod_exps, thetabar, n_days):
    """Sheet 1: per-epoch detail table."""
    from inputs.inputs import (N_LABS, L_AGE, GAMMA, LAMBDA_TOTAL,
                                DELTA_T, TAU_MAX)

    # Colour palette
    C_NAVY   = '1F3864'
    C_BLUE   = '2E75B6'
    C_DAY    = ['DEEAF1', 'FFF2CC', 'E2EFDA', 'FCE4D6']  # day0,day1,day2,day3
    C_DISP   = 'C6EFCE'   # light green  – dispatch rows
    C_EOD    = 'FCE4D6'   # salmon       – end-of-day rows
    C_ALT    = 'F2F2F2'   # grey         – alternating non-dispatch rows

    BRD  = _border()
    FMT2 = '0.00'

    ws = wb.active
    ws.title = 'Policy Simulation'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A5'

    # ── Row 1: title ──────────────────────────────────────────────────────────
    last_col = get_column_letter(4 + L_AGE * (3 + N_LABS * 2) + (1 + N_LABS) + 1 + 5)
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value     = f'POLICY SIMULATION REPORT  —  {n_days} Day(s) Under Greedy VFA Policy'
    c.font      = _hfont(13, True)
    c.fill      = _solid(C_NAVY)
    c.alignment = _center()
    ws.row_dimensions[1].height = 22

    # ── Row 2: parameter bar ──────────────────────────────────────────────────
    ws.merge_cells(f'A2:{last_col}2')
    c = ws['A2']
    c.value = (f'γ={GAMMA}   Λ={LAMBDA_TOTAL}/hr   '
               f'Δt=1/Λ={round(DELTA_T,4)} hr/epoch ({round(DELTA_T*60,1)} min)   '
               f'Epochs/day={int(TAU_MAX/DELTA_T)}   '
               f'L={L_AGE} age classes   P={N_LABS} labs   Seed fixed')
    c.font      = Font(name='Arial', size=9, italic=True, color='FFFFFF')
    c.fill      = _solid(C_BLUE)
    c.alignment = _center()
    ws.row_dimensions[2].height = 15

    # ── Rows 3–4: column headers ──────────────────────────────────────────────
    # Groups: (group_label, [(col_label, width), ...])
    groups = [
        ('Epoch Info',
         [('Step', 5), ('Day', 4), ('τ (hr)', 7), ('Event', 13)]),
        ('Arrivals',
         [(f'a{k+1}', 5) for k in range(L_AGE)]),
        ('Depot Inventory',
         [(f'a{k+1}', 5) for k in range(L_AGE)]),
    ]
    for p in range(N_LABS):
        groups.append((f'Lab {p+1} Inventory',
                       [(f'a{k+1}', 5) for k in range(L_AGE)]))
    for p in range(N_LABS):
        groups.append((f'Processed L{p+1}',
                       [(f'a{k+1}', 5) for k in range(L_AGE)]))
    groups.append(('Expiries (this epoch, EOD only)',
                   [('Depot', 6)] + [(f'L{p+1}', 5) for p in range(N_LABS)]))
    groups.append(('Decision', [('Action', 14)]))
    groups.append(('Costs',
                   [('c_disp', 7), ('c_hold', 7), ('c_exp', 7),
                    ('c_epoch', 7), ('CumCost', 9)]))

    col = 1
    col_meta = []   # (col_index, label)
    for g_label, sub in groups:
        span = len(sub)
        if span > 1:
            ws.merge_cells(start_row=3, start_column=col,
                           end_row=3,   end_column=col + span - 1)
        c = ws.cell(row=3, column=col, value=g_label)
        c.font      = _hfont(9, True)
        c.fill      = _solid(C_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        c.border    = BRD
        for s_label, w in sub:
            c4 = ws.cell(row=4, column=col, value=s_label)
            c4.font      = _hfont(8, True)
            c4.fill      = _solid(C_NAVY)
            c4.alignment = _center()
            c4.border    = BRD
            ws.column_dimensions[get_column_letter(col)].width = w
            col_meta.append((col, s_label))
            col += 1

    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 15

    # ── Data rows ─────────────────────────────────────────────────────────────
    cumcost   = 0.0
    n_cols    = len(col_meta)
    eod_row_indices = set()

    for ri, row in enumerate(rows):
        r      = 5 + ri
        day    = row['day']
        c_tot  = row['c_disp'] + row['c_hold'] + row['c_exp']
        cumcost += c_tot

        # detect EOD row
        is_eod = row['tau'] - DELTA_T <= 0

        # choose background
        if is_eod:
            bg = C_EOD
        elif row['action'] > 0:
            bg = C_DISP
        else:
            bg = C_DAY[day % len(C_DAY)] if ri % 2 == 0 else C_ALT

        # build value list matching col_meta order
        vals = (
            [row['step'], row['day'], row['tau'], row['event']]
            + row['arrive']
            + row['depot']
            + sum([row['labs'][p] for p in range(N_LABS)], [])
            + sum([row['processed'][p] for p in range(N_LABS)], [])
            + [row['eps_depot']] + row['eps_labs']
            + (['no dispatch'] if row['action'] == 0
               else [f'→ lab {row["action"]}'])
            + [row['c_disp'], row['c_hold'], row['c_exp'],
               round(c_tot, 4), round(cumcost, 4)]
        )

        bold_row = is_eod or row['action'] > 0
        for ci_idx, (ci, lbl) in enumerate(col_meta):
            val  = vals[ci_idx]
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill   = _solid(bg)
            cell.font   = _dfont(9, bold_row)
            cell.border = BRD
            if isinstance(val, float):
                cell.number_format = FMT2
                cell.alignment = _right()
            elif isinstance(val, int) and ci_idx > 3:
                cell.alignment = _center()
            elif ci_idx == 3:          # event
                cell.alignment = _left()
            elif ci_idx in (0, 1):
                cell.alignment = _center()
            elif ci_idx == 2:          # tau
                cell.number_format = '0.0'
                cell.alignment = _right()

        ws.row_dimensions[r].height = 13

    return ws


def _build_eod_sheet(wb, rows, eod_exps, n_days):
    """Sheet 2: end-of-day summary."""
    from inputs.inputs import N_LABS, C_EXP_DEPOT, C_EXP_LAB

    C_NAVY = '1F3864'
    C_BLUE = '2E75B6'
    C_DAY  = ['DEEAF1', 'FFF2CC', 'E2EFDA', 'FCE4D6']
    BRD    = _border()

    ws = wb.create_sheet('EOD Summary')
    ws.sheet_view.showGridLines = False

    last_col = get_column_letter(4 + N_LABS + 4)
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value     = 'End-of-Day Summary — Expiry & Cost Breakdown'
    c.font      = _hfont(12, True)
    c.fill      = _solid(C_NAVY)
    c.alignment = _center()
    ws.row_dimensions[1].height = 20

    headers = (['Day', 'Depot Exp.'] +
               [f'Lab {p+1} Exp.' for p in range(N_LABS)] +
               ['Total Exp.', 'Exp. Cost', 'Hold Cost', 'Disp Cost', 'Day Total'])
    widths  = [6, 10] + [10]*N_LABS + [10, 10, 10, 10, 11]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _hfont(9)
        c.fill      = _solid(C_BLUE)
        c.alignment = _center()
        c.border    = BRD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 16

    for day in range(n_days):
        exp      = eod_exps.get(day, {'eps_depot': 0, 'eps_labs': [0]*N_LABS})
        exp_d    = exp['eps_depot']
        exp_ls   = exp['eps_labs']
        exp_tot  = exp_d + sum(exp_ls)
        exp_cost = C_EXP_DEPOT * exp_d + C_EXP_LAB * sum(exp_ls)
        day_rows = [rw for rw in rows if rw['day'] == day]
        hold_tot = sum(rw['c_hold'] for rw in day_rows)
        disp_tot = sum(rw['c_disp'] for rw in day_rows)
        day_tot  = hold_tot + disp_tot + exp_cost

        vals = ([day, exp_d] + exp_ls +
                [exp_tot, round(exp_cost, 2),
                 round(hold_tot, 2), round(disp_tot, 2), round(day_tot, 2)])
        bg = C_DAY[day % len(C_DAY)]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=3 + day, column=ci, value=v)
            cell.fill      = _solid(bg)
            cell.font      = _dfont(9)
            cell.border    = BRD
            cell.alignment = _center()
            if isinstance(v, float):
                cell.number_format = '0.00'
        ws.row_dimensions[3 + day].height = 14

    return ws


def _build_theta_sheet(wb, thetabar, alp=None, mdp=None):
    """Sheet 3: VFA thetabar weights."""
    C_NAVY = '1F3864'
    C_BLUE = '2E75B6'
    BRD    = _border()

    ws = wb.create_sheet('VFA Parameters')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value     = 'Value Function Approximation — Final Thetabar Weights'
    c.font      = _hfont(12, True)
    c.fill      = _solid(C_NAVY)
    c.alignment = _center()
    ws.row_dimensions[1].height = 20

    hdrs   = ['Index', 'Feature', 'Description', 'θ̃ (normalized)', 'σ_b', 'θ (raw)']
    widths = [7, 16, 50, 16, 10, 12]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _hfont(9)
        c.fill      = _solid(C_BLUE)
        c.alignment = _center()
        c.border    = BRD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 16

    # Feature labels — MUST match the column order of ALP.phi() in classes/alp.py.
    # Keep this list in exact correspondence with phi(); slicing to _B guards the
    # length but NOT the order, so this is the single source of truth for names.
    from classes.alp import ALP as _ALP
    _B = _ALP.B
    _all_features = [
        ('const',         'Constant term (intercept)'),
        ('depot_age1',    'Depot age-1 holding  h_{0,1}·n_{0,1}'),
        ('depot_age2',    'Depot age-2 holding  h_{0,2}·n_{0,2}'),
        ('depot_age3',    'Depot age-3 holding  h_{0,3}·n_{0,3}'),
        ('lab_shortfall', 'Idle lab capacity  Σ_{p≥1} max(0, μ_p·τ − N_lab_p)'),
        ('imbalance',     'Lab backlog gap  |N_1/μ_1 − N_2/μ_2|'),
        ('exp_risk',      'Age-1 expiry risk  C_dep·n_{0,1} + C_lab·Σ_{p≥1} n_{p,1}'),
    ]
    if len(_all_features) < _B:
        raise ValueError(
            f"export label list has {len(_all_features)} names but ALP.B={_B}; "
            f"update _all_features in export_policy_excel.py to match alp.py phi()."
        )
    features = _all_features[:_B]

    # Normalization constants (if calibrated). θ passed in is in NORMALIZED
    # coordinates (what PSMD optimized). Report it alongside σ_b and the
    # de-normalized raw-coordinate weight θ_b = θ̃_b · cost_scale / σ_b
    # (θ_0 maps by cost_scale only). If not calibrated, σ_b=1, cost_scale=1.
    sigma_vec = np.asarray(getattr(alp, '_phi_scale', np.ones(_B)), float) if alp is not None else np.ones(_B)
    cscale    = float(getattr(mdp, 'cost_scale', 1.0)) if mdp is not None else 1.0

    for i, ((fname, fdesc), th) in enumerate(zip(features, thetabar)):
        r   = 3 + i
        bg  = 'EBF3FB' if i % 2 == 0 else 'FFFFFF'
        sg  = sigma_vec[i] if i < len(sigma_vec) else 1.0
        th_raw = (th * cscale) if i == 0 else (th * cscale / sg)
        row_vals = [i, fname, fdesc, round(float(th), 6),
                    round(float(sg), 4), round(float(th_raw), 6)]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.fill      = _solid(bg)
            cell.font      = _dfont(9)
            cell.border    = BRD
            cell.alignment = _right() if ci == 4 else _left()
            if ci == 4:
                cell.number_format = '0.000000'
        ws.row_dimensions[r].height = 14

    return ws


def _build_legend_sheet(wb):
    """Sheet 4: legend and model notes."""
    from inputs.inputs import (N_LABS, L_AGE, GAMMA, LAMBDA_TOTAL,
                                DELTA_T, TAU_MAX, C_EXP_DEPOT,
                                C_EXP_LAB, C_DISPATCH, H_HOLD,
                                LAMBDA_AGE, MU)

    C_NAVY = '1F3864'
    C_BLUE = '2E75B6'
    BRD    = _border()

    ws = wb.create_sheet('Legend')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 62

    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value     = 'Legend & Model Notes'
    c.font      = _hfont(12, True)
    c.fill      = _solid(C_NAVY)
    c.alignment = _center()
    ws.row_dimensions[1].height = 20

    items = [
        # (key, value) — key only with empty value = section header
        ('COLOUR CODING', ''),
        ('Blue rows',     'Day 0 epochs (no dispatch)'),
        ('Yellow rows',   'Day 1 epochs (no dispatch)'),
        ('Green rows',    'Epochs where a dispatch decision was made'),
        ('Salmon rows',   'Last epoch of each day (end-of-day transition)'),
        ('', ''),
        ('AGE CLASSES', ''),
        ('a1', '1 period remaining — expires tonight if not processed'),
        ('a2', '2 periods remaining'),
        ('a3', f'Freshest arrival — {L_AGE} periods remaining'),
        ('', ''),
        ('EVENTS', ''),
        ('arrive aK',   'Kit of age class K arrives at depot this epoch'),
        ('proc Lp',     'Lab p completes processing of one kit (oldest-first rule)'),
        ('dummy',       'No physical event — clock-tick only (uniformisation idle)'),
        ('', ''),
        ('KEY COLUMNS', ''),
        ('Step',        'Global epoch counter (0-indexed)'),
        ('τ (hr)',      'Remaining work hours in current day  [decrements by Δt each epoch]'),
        ('Arrivals',    'Kits arriving at depot this epoch, by age class'),
        ('Depot Inv.',  'Depot inventory before dispatch decision, by age class'),
        ('Lab Inv.',    'Lab inventory before dispatch decision, by age class'),
        ('Processed',   'Kits removed from lab this epoch (processing completion)'),
        ('Expiries',    'Realised age-1 kits expiring THIS epoch (only nonzero on the '
                        'end-of-day epoch; the state carries no expiry counter — see '
                        'Section 3.3)'),
        ('Action',      '"no dispatch" or "→ lab p" — greedy VFA decision'),
        ('c_disp',      'Dispatch cost = C_p if dispatched to lab p, else 0'),
        ('c_hold',      'Holding cost this epoch = Σ_{p,a} h_{p,a}·n_{p,a} / Λ'),
        ('c_exp',       'Expiry cost = C_exp_depot·ε_0 + C_exp_lab·Σ_{p≥1} ε_p, realised '
                        'value on the end-of-day epoch only (mdp.cost() uses the exact '
                        'closed-form expectation of this instead)'),
        ('c_epoch',     'Total cost this epoch = c_disp + c_hold + c_exp'),
        ('CumCost',     'Running cumulative cost (undiscounted)'),
        ('', ''),
        ('MODEL PARAMETERS', ''),
        ('γ (gamma)',        str(GAMMA)),
        ('Λ',               f'{LAMBDA_TOTAL} /hr  =  λ + Σμ_p'),
        ('Δt (epoch)',       f'{round(DELTA_T, 6)} hr  =  1/Λ  ≈  {round(DELTA_T*60, 2)} minutes'),
        ('Epochs / day',    str(int(TAU_MAX / DELTA_T))),
        ('λ_age (arr.rates)',str(list(LAMBDA_AGE)) + ' /hr'),
        ('μ_p (proc.rates)', str(list(MU)) + ' /hr'),
        ('C_dispatch',       str(list(C_DISPATCH))),
        ('C_exp_depot',      str(C_EXP_DEPOT)),
        ('C_exp_lab',        str(C_EXP_LAB)),
        ('TAU_MAX',          f'{TAU_MAX} hours/day'),
        ('Horizon',          'Infinite, discounted, stationary (no terminal day)'),
    ]

    for ri, (k, v) in enumerate(items, 2):
        is_section = (v == '' and k != '')
        ck = ws.cell(row=ri, column=1, value=k)
        cv = ws.cell(row=ri, column=2, value=v)
        if is_section:
            for cx in (ck, cv):
                cx.font      = _hfont(9, True)
                cx.fill      = _solid(C_BLUE)
                cx.alignment = _left()
                cx.border    = BRD
        else:
            ck.font = _dfont(9, bold=True)
            cv.font = _dfont(9)
            bg = 'EBF3FB' if ri % 2 == 0 else 'FFFFFF'
            ck.fill = cv.fill = _solid(bg)
            ck.alignment = _left()
            cv.alignment = Alignment(horizontal='left', vertical='center',
                                     wrap_text=True)
            ck.border = cv.border = BRD
        ws.row_dimensions[ri].height = 14

    return ws


# ── main entry point ─────────────────────────────────────────────────────────

def export_policy_excel(mdp, alp, thetabar,
                        n_days=5, seed=7,
                        filename='dispatching_policy_report.xlsx'):
    """
    Run a greedy trajectory and export a 4-sheet Excel report.

    Parameters
    ----------
    mdp       : MDP instance
    alp       : ALP instance
    thetabar  : (B,) final VFA weight vector from run_psmd()
    n_days    : int   — number of days to simulate  (default 2)
    seed      : int   — random seed for reproducibility  (default 7)
    filename  : str   — output file path  (default 'dispatching_policy_report.xlsx')
    """
    print(f'\nRunning policy simulation ({n_days} days, seed={seed})...')
    rows, eod_exps = _run_simulation(mdp, alp, thetabar, n_days, seed)
    print(f'  {len(rows)} epochs collected across {n_days} days.')

    wb = Workbook()
    _build_simulation_sheet(wb, rows, eod_exps, thetabar, n_days)
    _build_eod_sheet(wb, rows, eod_exps, n_days)
    _build_theta_sheet(wb, thetabar, alp=alp, mdp=mdp)
    _build_legend_sheet(wb)

    wb.save(filename)
    print(f'  Excel report saved → {filename}')
    return filename